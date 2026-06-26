from flask import Flask, request, jsonify, render_template_string, send_file, session, redirect, url_for
import os
import sys
import re
import csv
from functools import wraps
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix

# Carrega variáveis de ambiente do arquivo .env
from dotenv import load_dotenv
load_dotenv()

# Add current directory to path before imports so local modules are found
base_dir = os.path.dirname(os.path.abspath(__file__))
if base_dir not in sys.path:
    sys.path.insert(0, base_dir)

from admin_db import (
    init_db, get_modulos, get_conteudos_modulo, get_dados_conteudo,
    criar_conteudo, atualizar_dados_conteudo, atualizar_tag_conteudo, atualizar_arquivo_conteudo, deletar_conteudo,
    buscar_conteudo_modulo, buscar_conteudos_modulo_por_query
)

# Import logic from chatgpt_integration
CHATGPT_HABILITADO = False
try:
    from chatgpt_integration import (
        analisar_com_chatgpt,
        interpretar_pergunta,
        gerar_resposta_com_fallback,
        CHATGPT_HABILITADO,
        validar_chave_api
    )
    print("Import from chatgpt_integration successful")
except ImportError as e:
    print(f"Import failed: {e}")
    CHATGPT_HABILITADO = False
    def analisar_com_chatgpt(*args, **kwargs):
        return None
    def interpretar_pergunta(*args, **kwargs):
        return None
    def gerar_resposta_com_fallback(pergunta, base=None, fallback_fn=None):
        return "Resposta padrão: Integração ChatGPT não disponível."
    def validar_chave_api():
        return False

# Load knowledge bases
import json

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
knowledge_bases = {}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def load_knowledge_bases():
    # Bases gerais e orçamento
    base_files = [
        'ai_base_geral.json',
        'ai_base_orcamento.json', 
        'ai_base_qualidade.json',
        'ai_base_orcamento_sense.json',
        'ai_base_orcamento_corporate.json',
        'ai_base_orcamento_passeio.json'
    ]
    
    # Adicionar todas as bases de qualidade dos PES
    import glob
    pes_files = glob.glob(os.path.join(BASE_DIR, 'ai_base_qualidade_*.json'))
    base_files.extend([os.path.basename(f) for f in pes_files])
    
    # Adicionar bases de texto
    texto_files = glob.glob(os.path.join(BASE_DIR, 'ai_base_texto_*.json'))
    base_files.extend([os.path.basename(f) for f in texto_files])
    
    for file in base_files:
        path = os.path.join(BASE_DIR, file)
        if os.path.exists(path):
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    # Normalize to list of {'question': str, 'answer': str}
                    knowledge_bases[file] = normalize_base(data)
                print(f"Loaded knowledge base: {file}")
            except Exception as e:
                print(f"Error loading {file}: {e}")

def normalize_base(data):
    """Normalize different JSON structures to a list of {'question': str, 'answer': str}"""
    normalized = []
    
    if isinstance(data, dict):
        # Check for 'qa' structure
        if 'qa' in data and isinstance(data['qa'], list):
            for item in data['qa']:
                if isinstance(item, dict) and 'q' in item and 'a' in item:
                    normalized.append({
                        'question': item['q'],
                        'answer': item['a']
                    })
        # Check for 'textos' structure
        elif 'textos' in data and isinstance(data['textos'], list):
            for item in data['textos']:
                if isinstance(item, dict) and 'titulo' in item and 'conteudo' in item:
                    normalized.append({
                        'question': item['titulo'],
                        'answer': item['conteudo']
                    })
        # Check for question: answer dict
        else:
            for q, a in data.items():
                if isinstance(a, str):
                    normalized.append({
                        'question': q,
                        'answer': a
                    })
    
    return normalized


def load_spreadsheet_summary(file_path, max_rows=30):
    ext = os.path.splitext(file_path)[1].lower().strip('.')
    if ext in {'xlsx', 'xls'}:
        try:
            import openpyxl
        except ImportError:
            return None
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        headers = []
        if sheet.max_row >= 1:
            for col in range(1, sheet.max_column + 1):
                value = sheet.cell(row=1, column=col).value
                headers.append(str(value).strip() if value is not None else f'col{col}')
        rows = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column, values_only=True):
            if not any(cell is not None for cell in row):
                continue
            row_data = []
            for idx, cell in enumerate(row):
                header = headers[idx] if idx < len(headers) else f'col{idx+1}'
                row_data.append(f"{header}: {'' if cell is None else str(cell)}")
            rows.append('; '.join(row_data))
            if len(rows) >= max_rows:
                break
        return '\n'.join(rows)
    elif ext == 'csv':
        try:
            with open(file_path, newline='', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                headers = next(reader, None)
                rows = []
                for row in reader:
                    if not any(cell.strip() for cell in row if isinstance(cell, str)):
                        continue
                    row_data = []
                    for idx, cell in enumerate(row):
                        header = headers[idx] if headers and idx < len(headers) else f'col{idx+1}'
                        row_data.append(f"{header}: {cell.strip() if isinstance(cell, str) else cell}")
                    rows.append('; '.join(row_data))
                    if len(rows) >= max_rows:
                        break
            return '\n'.join(rows)
        except Exception:
            return None
    return None

load_knowledge_bases()

def extract_relevant_sentences(text, keywords, max_sentences=5):
    sentences = re.split(r'(?<=[.!?])\s+', text)
    selected = [s.strip() for s in sentences if any(k in s.lower() for k in keywords)]
    return ' '.join(selected[:max_sentences]) if selected else None

def extract_section_by_heading(text, heading_keywords):
    # Split by common section markers
    sections = re.split(r'(Para a Instalação de|No procedimento de|Instalação de|Instalação da)', text)
    for i, section in enumerate(sections):
        if any(k in section.lower() for k in heading_keywords):
            # Return this section content (between markers)
            return section.strip()
    return None


def clean_direct_response(text):
    if not isinstance(text, str):
        return text
    text = text.strip()
    text = re.sub(r'^(ol[áa]s?\b[:,]?\s*|olá\b[:,]?\s*|oi\b[:,]?\s*|bom dia\b[:,]?\s*|boa tarde\b[:,]?\s*|boa noite\b[:,]?\s*)', '', text, flags=re.I)
    text = re.sub(r'^(desculpe[\s\S]*?[\.!?]\s*)', '', text, flags=re.I)
    text = re.sub(r'^(infelizmente[\s\S]*?[\.!?]\s*)', '', text, flags=re.I)
    text = re.sub(r'^(claro[\s\S]*?[\.!?]\s*)', '', text, flags=re.I)
    text = re.sub(r'^(o\s+procedimento\s+de\s+[\s\w\-çãõáéíóúâêô]+?\s+estabelece(?:\s+como)?\s+)', '', text, flags=re.I)
    text = re.sub(r'^(o\s+(?:procedimento|documento)\s+[\s\w\-çãõáéíóúâêô]+?\s+define\s+)', '', text, flags=re.I)
    text = re.sub(r'^(o\s+(?:procedimento|documento)\s+[\s\w\-çãõáéíóúâêô]+?\s+de\s+.*?\s+estabelece\s+)', '', text, flags=re.I)
    text = re.sub(r'(\s*(Se precisar|Caso queira|Se quiser|Se tiver|Se precisar).*?$)', '', text, flags=re.I|re.S)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

# Função de busca direta nas bases de qualidade (PES)
def obter_resposta_base(question, base):
    if not base or not isinstance(base, list):
        return None

    question_lower = question.lower()

    # score por similaridade básica para evitar retorno muito genérico
    best_match = None
    best_score = 0
    question_words = [w for w in re.findall(r"\w+", question_lower) if len(w) > 2]

    for item in base:
        if not isinstance(item, dict) or 'question' not in item or 'answer' not in item:
            continue
        
        q = item['question'].lower()
        a = item['answer']
        
        # Match direto por termo
        score = sum(1 for w in question_words if w in q or w in a.lower())
        
        # +1 quando pergunta e item claramente relacionados a metais/pes/instalação
        if 'instalação de metais' in question_lower or 'instalacao de metais' in question_lower:
            if 'instalação de metais' in q or 'instalacao de metais' in q:
                score += 5
        if 'torneira' in question_lower and 'pared' in question_lower:
            if 'torneira' in q and 'pared' in q:
                score += 7
            if 'torneira' in a.lower() and 'pared' in a.lower():
                score += 5

        if score > best_score:
            best_score = score
            best_match = item

    if not best_match or best_score < 1:  # Lower threshold for better matching
        return None
    
    answer = best_match['answer']
    
    # Extract relevant sections for specific keywords
    if 'requsit' in question_lower or 'pre' in question_lower:
        # Extract prerequisites directly
        found = re.findall(r'(?:(?:pré[- ]requisitos?|requisitos)[^.!?]*[.!?])', answer, flags=re.I)
        if found:
            return ' '.join([clean_direct_response(s) for s in found])
        sentences = re.split(r'(?<=[.!?])\s+', answer)
        found = [s for s in sentences if 'requisit' in s.lower() or 'pré' in s.lower() or 'requisitos' in s.lower()]
        if found:
            return ' '.join([clean_direct_response(s) for s in found])
        paragraphs = answer.split('\n\n')
        return '\n\n'.join([clean_direct_response(p) for p in paragraphs[:2]]) if paragraphs else clean_direct_response(answer)

    # Specific extractions for common queries
    if 'sifão' in question_lower or 'sifoes' in question_lower or 'escoamento' in question_lower:
        section = extract_section_by_heading(answer, ['válvula de escoamento', 'sifões'])
        if section:
            sentences = re.split(r'(?<=[.!?])\s+', section)
            passos = [s.strip() for s in sentences if any(k in s.lower() for k in ['verificar', 'posicionar', 'rosquear', 'selecionar', 'montar', 'substituir', 'evitar'])]
            if passos:
                return ' '.join(passos[:10])
            return section.strip()

    if 'engate' in question_lower:
        section = extract_section_by_heading(answer, ['engate'])
        if section:
            sentences = re.split(r'(?<=[.!?])\s+', section)
            passos = [s.strip() for s in sentences if any(k in s.lower() for k in ['rosquear', 'observando', 'ligamento', 'passar', 'instalar'])]
            if passos:
                return ' '.join(passos[:10])
            return section.strip()

    if 'passo a passo' in question_lower or 'passo-a-passo' in question_lower or 'procedimento' in question_lower:
        relevant = extract_relevant_sentences(answer, ['passo', 'procedimento', 'verificar', 'instala', 'torneira'])
        if relevant:
            return relevant

    # For torneira de parede
    if 'torneira' in question_lower and 'pared' in question_lower:
        section = extract_section_by_heading(answer, ['torneira de parede'])
        if section:
            sentences = re.split(r'(?<=[.!?])\s+', section)
            passos = [s.strip() for s in sentences if any(k in s.lower() for k in ['verificar', 'encaixar', 'aplicar', 'instalar', 'rosquear', 'checar'])]
            if passos:
                return ' '.join(passos[:10])
            return section.strip()

    # For misturador
    if 'misturador' in question_lower:
        section = extract_section_by_heading(answer, ['misturador'])
        if section:
            sentences = re.split(r'(?<=[.!?])\s+', section)
            passos = [s.strip() for s in sentences if any(k in s.lower() for k in ['verificar', 'rosquear', 'posicionar', 'passar', 'instalar', 'testar', 'checar'])]
            if passos:
                return ' '.join(passos[:10])
            return section.strip()

    # For torneira de bancada
    if 'torneira' in question_lower and 'bancada' in question_lower:
        section = extract_section_by_heading(answer, ['torneira de bancada'])
        if section:
            sentences = re.split(r'(?<=[.!?])\s+', section)
            passos = [s.strip() for s in sentences if any(k in s.lower() for k in ['verificar', 'encaixar', 'posicionar', 'rosquear', 'não apertar'])]
            if passos:
                return ' '.join(passos[:10])
            return section.strip()

    # Default: return relevant sentences or full answer
    sentences = re.split(r'(?<=[.!?])\s+', answer)
    matches = [s.strip() for s in sentences if any(w in s.lower() for w in question_words[:8])]
    if matches:
        return ' '.join(matches[:5])
    
    return answer

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1, x_prefix=1)
app.secret_key = os.getenv('FLASK_SECRET_KEY', 'zia_admin_opus_2024_secure_key')  # Altere para uma chave segura em produção

# Configurar limites para melhor desempenho em mobile
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB limite para uploads
app.config['JSON_SORT_KEYS'] = False
app.config['JSONIFY_PRETTYPRINT_REGULAR'] = False
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 3600

# Timeout para requisições longas (em segundos)
import signal
REQUEST_TIMEOUT = 60

# Inicializa o banco de dados do painel administrativo
init_db()

# Credenciais de admin (alterar em produção)
ADMIN_USERNAME = 'admin'
ADMIN_PASSWORD = 'opus123'  # Altere para uma senha segura

def login_required(f):
    """Decorator para proteger rotas de admin."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function


ADMIN_LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ZIA Admin Login</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: Arial, sans-serif; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; display: flex; align-items: center; justify-content: center; }
        .login-container { background: white; padding: 40px; border-radius: 10px; box-shadow: 0 10px 25px rgba(0,0,0,0.3); width: 100%; max-width: 400px; }
        h1 { text-align: center; color: #333; margin-bottom: 30px; font-size: 24px; }
        .form-group { margin-bottom: 20px; }
        label { display: block; margin-bottom: 8px; color: #555; font-weight: bold; }
        input[type="text"], input[type="password"] { width: 100%; padding: 12px; border: 1px solid #ddd; border-radius: 5px; font-size: 14px; }
        input[type="text"]:focus, input[type="password"]:focus { outline: none; border-color: #667eea; }
        button { width: 100%; padding: 12px; background: #667eea; color: white; border: none; border-radius: 5px; font-size: 16px; cursor: pointer; font-weight: bold; transition: 0.3s; }
        button:hover { background: #5568d3; }
        .erro { color: #d32f2f; background: #ffebee; padding: 12px; border-radius: 5px; margin-bottom: 20px; text-align: center; }
        .info { color: #666; text-align: center; margin-top: 20px; font-size: 12px; }
    </style>
</head>
<body>
    <div class="login-container">
        <h1>ZIA Admin</h1>
        {% if erro %}<div class="erro">{{ erro }}</div>{% endif %}
        <form method="POST">
            <div class="form-group">
                <label for="username">Usuário:</label>
                <input type="text" id="username" name="username" required autofocus>
            </div>
            <div class="form-group">
                <label for="password">Senha:</label>
                <input type="password" id="password" name="password" required>
            </div>
            <button type="submit">Entrar</button>
        </form>
        <div class="info">Painel de Administração ZIA</div>
    </div>
</body>
</html>
"""

ADMIN_PANEL_TEMPLATE = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ZIA Admin Panel</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: Arial, sans-serif; background: #f5f5f5; }
        .container { display: flex; min-height: 100vh; }
        .sidebar { width: 280px; background: #333; color: white; padding: 20px; overflow-y: auto; }
        .sidebar h2 { margin-bottom: 20px; font-size: 16px; color: #667eea; }
        .modulo-btn { display: block; width: 100%; padding: 12px 15px; margin-bottom: 8px; background: #555; color: white; border: none; border-radius: 5px; cursor: pointer; text-align: left; transition: 0.3s; font-size: 14px; }
        .modulo-btn:hover, .modulo-btn.active { background: #667eea; }
        .logout { position: absolute; bottom: 20px; left: 20px; right: 20px; padding: 10px; background: #d32f2f; color: white; border: none; border-radius: 5px; cursor: pointer; }
        .main-content { flex: 1; padding: 30px; overflow-y: auto; }
        .header { margin-bottom: 30px; }
        h1 { color: #333; font-size: 28px; }
        .breadcrumb { color: #666; font-size: 14px; margin-top: 10px; }
        .panel { background: white; border-radius: 10px; padding: 25px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); margin-bottom: 20px; }
        .panel h2 { color: #333; margin-bottom: 20px; border-bottom: 2px solid #667eea; padding-bottom: 10px; }
        .form-group { margin-bottom: 15px; }
        label { display: block; margin-bottom: 8px; color: #555; font-weight: bold; font-size: 14px; }
        input[type="text"], textarea { width: 100%; padding: 12px; border: 1px solid #ddd; border-radius: 5px; font-family: Arial; font-size: 14px; }
        textarea { resize: vertical; min-height: 300px; font-family: monospace; }
        textarea:focus { outline: none; border-color: #667eea; box-shadow: 0 0 5px rgba(102,126,234,0.5); }
        button { padding: 10px 20px; background: #667eea; color: white; border: none; border-radius: 5px; cursor: pointer; font-weight: bold; transition: 0.3s; font-size: 14px; }
        button:hover { background: #5568d3; }
        .btn-delete { background: #d32f2f; }
        .btn-delete:hover { background: #b71c1c; }
        .conteudo-list { margin-top: 15px; }
        .conteudo-item { background: #f9f9f9; border-left: 4px solid #667eea; padding: 15px; margin-bottom: 10px; border-radius: 5px; cursor: pointer; transition: 0.3s; }
        .conteudo-item:hover, .conteudo-item.ativo { background: #e8eef7; border-left-color: #5568d3; }
        .conteudo-nome { font-weight: bold; color: #333; }
        .conteudo-descricao { color: #666; font-size: 12px; margin-top: 5px; }
        .mensagem { padding: 15px; border-radius: 5px; margin-bottom: 15px; text-align: center; font-weight: bold; }
        .sucesso { background: #c8e6c9; color: #2e7d32; }
        .erro { background: #ffcdd2; color: #d32f2f; }
        .status-text { color: #666; font-size: 12px; margin-top: 5px; }
        .btn-small { padding: 6px 12px; font-size: 12px; }
    </style>
</head>
<body>
    <div class="container">
        <div class="sidebar">
            <h2>📚 Módulos</h2>
            {% for modulo in modulos %}
                <button class="modulo-btn" onclick="selecionarModulo({{ modulo.id }}, this)">
                    {{ modulo.nome|replace('_', ' ')|title }}
                </button>
            {% endfor %}
            <a href="/admin/logout" class="logout">🚪 Sair</a>
        </div>
        <div class="main-content">
            <div class="header">
                <h1>Gerenciar Conteúdo ZIA</h1>
                <div class="breadcrumb">Módulo: <span id="modulo-nome">Selecione um módulo</span></div>
            </div>
            
            <div id="mensagem"></div>
            
            <div class="panel">
                <h2>Conteúdos Disponíveis</h2>
                <div id="conteudo-list" class="conteudo-list">
                    <p style="color: #999; text-align: center;">Selecione um módulo para ver conteúdos</p>
                </div>
            </div>
            
            <div class="panel">
                <h2>Editar Conteúdo</h2>
                <div id="editor" style="display: none;">
                    <div class="form-group">
                        <label>Nome do Conteúdo:</label>
                        <input type="text" id="conteudo-nome-label" disabled style="background: #f5f5f5;">
                    </div>
                    <div class="form-group">
                        <label>Tag:</label>
                        <input type="text" id="conteudo-tag" placeholder="SENSE, corporate, passeio, etc.">
                    </div>
                    <div class="form-group">
                        <label>Arquivo de dados:</label>
                        <div style="display:flex; gap:10px; align-items:center;">
                            <input type="text" id="conteudo-file-name" disabled style="background: #f5f5f5; flex:1;" placeholder="Nenhum arquivo anexado">
                            <button type="button" onclick="selecionarArquivo()">📎 Anexar</button>
                            <button type="button" class="btn-delete btn-small" onclick="removerArquivo()">✕ Remover</button>
                        </div>
                        <input type="file" id="conteudo-file-input" accept=".xlsx,.xls,.csv" style="display:none;">
                    </div>
                    <div class="form-group">
                        <label>Dados/Conteúdo (texto com informações):</label>
                        <textarea id="conteudo-texto" placeholder="Cole aqui os dados e informações do conteúdo..."></textarea>
                        <div class="status-text">Última atualização: <span id="status-atualizado">-</span></div>
                    </div>
                    <button onclick="salvarConteudo()">💾 Salvar Dados</button>
                    <button class="btn-delete btn-small" onclick="deletarConteudo()">🗑️ Deletar Conteúdo</button>
                </div>
                <p id="editor-placeholder" style="color: #999; text-align: center;">Selecione um conteúdo para editar</p>
            </div>
        </div>
    </div>

    <script>
        let moduloAtual = null;
        let conteudoAtual = null;

        function selecionarModulo(moduloId, elemento) {
            moduloAtual = moduloId;
            document.querySelectorAll('.modulo-btn').forEach(btn => btn.classList.remove('active'));
            elemento.classList.add('active');
            
            carregarConteudosModulo(moduloId);
            limparEditor();
        }

        function carregarConteudosModulo(moduloId) {
            fetch(`/admin/conteudos/${moduloId}`)
                .then(r => r.json())
                .then(data => {
                    const lista = document.getElementById('conteudo-list');
                    const nomeSpan = document.getElementById('modulo-nome');
                    
                    if (!data.conteudos || data.conteudos.length === 0) {
                        lista.innerHTML = '<p style="color: #999;">Nenhum conteúdo disponível neste módulo.</p>';
                        return;
                    }
                    
                    nomeSpan.textContent = document.querySelector('.modulo-btn.active')?.textContent || 'Desconhecido';
                    lista.innerHTML = data.conteudos.map(item => `
                        <div class="conteudo-item" onclick="selecionarConteudo(${item.id}, '${item.nome}', this)">
                            <div class="conteudo-nome">${item.nome.replace(/_/g, ' ').toUpperCase()}</div>
                            <div class="conteudo-descricao">
                                ${item.descricao || 'Sem descrição'}
                                ${item.tag ? ' | Tag: ' + item.tag.toUpperCase() : ''}
                                ${item.arquivo ? ' | Arquivo: ' + item.arquivo.replace(/^conteudo_\\d+_/, '') : ''}
                            </div>
                        </div>
                    `).join('');
                });
        }

        function selecionarConteudo(conteudoId, conteudoNome, elemento) {
            conteudoAtual = conteudoId;
            document.querySelectorAll('.conteudo-item').forEach(item => item.classList.remove('ativo'));
            elemento.classList.add('ativo');
            
            document.getElementById('conteudo-nome-label').value = conteudoNome;
            document.getElementById('editor').style.display = 'block';
            document.getElementById('editor-placeholder').style.display = 'none';
            
            carregarDadosConteudo(conteudoId);
        }

        function carregarDadosConteudo(conteudoId) {
            fetch(`/admin/dados/${conteudoId}`)
                .then(r => r.json())
                .then(data => {
                    document.getElementById('conteudo-texto').value = data.texto || '';
                    document.getElementById('conteudo-tag').value = data.tag || '';
                    document.getElementById('conteudo-file-name').value = data.arquivo ? data.arquivo.replace(/^conteudo_\d+_/, '') : '';
                    document.getElementById('status-atualizado').textContent = new Date().toLocaleString('pt-BR');
                });
        }

        function selecionarArquivo() {
            if (!conteudoAtual) {
                mostrarMensagem('Selecione um conteúdo primeiro', 'erro');
                return;
            }
            document.getElementById('conteudo-file-input').value = '';
            document.getElementById('conteudo-file-input').click();
        }

        document.getElementById('conteudo-file-input').addEventListener('change', function(e) {
            const file = e.target.files[0];
            if (!file) return;
            const allowedExtensions = ['xlsx', 'xls', 'csv'];
            const extension = file.name.split('.').pop().toLowerCase();
            if (!allowedExtensions.includes(extension)) {
                mostrarMensagem('Formato não suportado. Use XLSX, XLS ou CSV.', 'erro');
                return;
            }

            const formData = new FormData();
            formData.append('arquivo', file);

            fetch(`/admin/arquivo/${conteudoAtual}`, {
                method: 'POST',
                body: formData
            })
            .then(r => r.json())
            .then(data => {
                if (data.sucesso) {
                    mostrarMensagem('Arquivo anexado com sucesso!', 'sucesso');
                    document.getElementById('conteudo-file-name').value = data.nome || file.name;
                } else {
                    mostrarMensagem('Erro: ' + data.mensagem, 'erro');
                }
            })
            .catch(err => {
                mostrarMensagem('Erro ao enviar arquivo.', 'erro');
                console.error(err);
            });
        });

        function removerArquivo() {
            if (!conteudoAtual) {
                mostrarMensagem('Selecione um conteúdo primeiro', 'erro');
                return;
            }
            fetch(`/admin/arquivo/${conteudoAtual}`, {
                method: 'POST',
                headers: { 'Content-Type': 'application/x-www-form-urlencoded' },
                body: 'remove=1'
            })
            .then(r => r.json())
            .then(data => {
                if (data.sucesso) {
                    mostrarMensagem('Arquivo removido com sucesso.', 'sucesso');
                    document.getElementById('conteudo-file-name').value = '';
                } else {
                    mostrarMensagem('Erro: ' + data.mensagem, 'erro');
                }
            })
            .catch(err => {
                mostrarMensagem('Erro ao remover arquivo.', 'erro');
                console.error(err);
            });
        }

        function salvarConteudo() {
            if (!conteudoAtual) {
                mostrarMensagem('Selecione um conteúdo primeiro', 'erro');
                return;
            }
            
            const texto = document.getElementById('conteudo-texto').value.trim();
            const tag = document.getElementById('conteudo-tag').value.trim();
            if (!texto) {
                mostrarMensagem('Adicione algum conteúdo antes de salvar', 'erro');
                return;
            }
            
            fetch(`/admin/dados/${conteudoAtual}`, {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ texto, tag })
            })
            .then(r => r.json())
            .then(data => {
                if (data.sucesso) {
                    mostrarMensagem('✅ Dados salvos com sucesso!', 'sucesso');
                    document.getElementById('status-atualizado').textContent = new Date().toLocaleString('pt-BR');
                } else {
                    mostrarMensagem('❌ Erro: ' + data.mensagem, 'erro');
                }
            });
        }

        function deletarConteudo() {
            if (!conteudoAtual) return;
            if (!confirm('Tem certeza que quer deletar este conteúdo?')) return;
            
            fetch(`/admin/deletar-conteudo/${conteudoAtual}`, { method: 'POST' })
                .then(r => r.json())
                .then(data => {
                    if (data.sucesso) {
                        mostrarMensagem('Conteúdo deletado com sucesso', 'sucesso');
                        limparEditor();
                        if (moduloAtual) carregarConteudosModulo(moduloAtual);
                    } else {
                        mostrarMensagem('Erro ao deletar: ' + data.mensagem, 'erro');
                    }
                });
        }

        function limparEditor() {
            conteudoAtual = null;
            document.getElementById('editor').style.display = 'none';
            document.getElementById('editor-placeholder').style.display = 'block';
            document.getElementById('conteudo-nome-label').value = '';
            document.getElementById('conteudo-texto').value = '';
            document.querySelectorAll('.conteudo-item').forEach(item => item.classList.remove('ativo'));
        }

        function mostrarMensagem(texto, tipo) {
            const el = document.getElementById('mensagem');
            el.className = 'mensagem ' + tipo;
            el.textContent = texto;
            setTimeout(() => el.textContent = '', 3000);
        }
    </script>
</body>
</html>
"""

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ZIA Web</title>
    <link rel="icon" type="image/png" href="/favicon">
    <link rel="shortcut icon" type="image/png" href="/favicon">
    <style>
        body {
            font-family: Arial, sans-serif;
            max-width: 100%;
            width: 100%;
            margin: 0;
            padding: 0;
            background-color: white;
        }
        .header {
            text-align: center;
            padding: 10px 20px;
            background: white;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            margin-bottom: 0;
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            width: 100%;
            z-index: 1001;
        }

        .header-logo {
            height: 50px;
            max-width: 90%;
            object-fit: contain;
        }

        .context-selector {
            display: flex;
            flex-direction: column;
            gap: 10px;
            margin-bottom: 20px;
            width: 100%;
        }
        .context-btn {
            width: 100%;
            max-width: 300px;
            padding: 10px 14px;
            border: 1px solid #007bff;
            background: white;
            color: #007bff;
            border-radius: 5px;
            cursor: pointer;
            text-align: center;
            white-space: normal;
            word-wrap: break-word;
            font-size: 13px;
        }
        .context-btn.active {
            background: #007bff;
            color: white;
        }
        .header {
            text-align: center;
            padding: 10px 20px;
            background: white;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            margin-bottom: 0;
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            width: 100%;
            z-index: 1001;
        }

        .header h1 {
            margin: 0;
            font-size: 1.5rem;
        }

        .chat-container {
            background: white;
            border-radius: 0;
            padding: 20px;
            box-shadow: none;
            height: calc(100vh - 160px);
            min-height: 520px;
            overflow-y: auto;
            margin-top: 60px; /* espaço para header fixo */
            margin-bottom: 80px; /* espaço para input fixo */
        }

        .input-container,
        .file-input {
            position: fixed;
            left: 0;
            right: 0;
            width: 100%;
            background: white;
            padding: 10px 20px;
            z-index: 1000;
        }

        .file-input {
            bottom: 80px;
        }

        .input-container {
            bottom: 0;
            display: flex;
            gap: 10px;
            align-items: center;
            padding: 15px 20px;
            justify-content: flex-end; /* botão no canto direito */
        }

        input[type="text"] {
            flex: 1;
            padding: 15px 20px; /* aumentado o padding */
            border: 1px solid #ddd;
            border-radius: 25px;
            font-size: 14px; /* fonte reduzida */
            min-width: 0;
            min-height: 20px; /* altura mínima maior */
        }

        button {
            padding: 10px 20px;
            background: #007bff;
            color: white;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            font-size: 16px;
        }

        .send-button {
            width: 50px;
            height: 50px;
            background: transparent;
            border: none;
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: center;
            transition: all 0.3s ease;
            flex-shrink: 0;
            position: relative;
        }

        .send-button::before {
            content: '';
            position: absolute;
            width: 0;
            height: 0;
            border-left: 15px solid black;
            border-top: 12px solid transparent;
            border-bottom: 12px solid transparent;
            left: 14%; /* movido mais para a esquerda */
            top: 50%;
            transform: translate(-50%, -50%);
        }

        .send-button.stop-typing::before {
            width: 18px;
            height: 18px;
            border: none;
            background: black;
        }

        .chat-container {
            margin-bottom: 140px; /* espaço para controles fixos */
        }
        .message {
            margin-bottom: 15px;
            padding: 10px;
            border-radius: 10px;
            font-size: 15px;
            line-height: 1.4;
        }
        .user-message {
            background: #F5FAFC;
            color: #333;
            text-align: left;
        }
        .bot-message {
            background: transparent;
            color: black;
        }
        .message-prefix {
            color: #8B0000;
            font-weight: bold;
        }
        .typing-indicator {
            display: inline-block;
        }
        .typing-indicator::after {
            content: '...';
            animation: typing 1.5s infinite;
        }
        @keyframes typing {
            0%, 20% { content: ''; }
            40% { content: '.'; }
            60% { content: '..'; }
            80%, 100% { content: '...'; }
        }
        button:hover:not(.send-button) {
            background: #0056b3;
        }

        .send-button:hover {
            background: transparent;
        }
        .input-container {
            position: fixed;
            bottom: 0;
            left: 0;
            right: 0;
            display: flex;
            align-items: center;
            width: 100%;
            padding: 10px 10px 15px; /* reduzido padding horizontal */
            box-sizing: border-box;
            z-index: 1000;
            background: white; /* fundo para cobrir conteúdo */
        }

        .input-wrapper {
            position: relative;
            flex: 1;
            min-width: 0;
            display: flex;
            align-items: center;
        }

        .input-wrapper input[type="text"] {
            width: 100%;
            padding: 18px 120px 18px 16px; /* espaço para botão enviar e clip */
            border: 1px solid #ddd;
            border-radius: 25px;
            font-size: 16px;
            min-height: 52px;
            box-sizing: border-box;
            transition: padding-left 0.3s ease;
            flex: 1;
        }

        .input-wrapper.has-attachment input[type="text"] {
            padding-left: 120px; /* espaço para a pré-visualização */
        }

        .attachment-preview {
            position: absolute;
            left: 10px;
            top: 50%;
            transform: translateY(-50%);
            max-width: 100px;
            display: flex;
            opacity: 1 !important;
            pointer-events: auto !important;
            align-items: center;
            gap: 6px;
            font-size: 14px;
            color: #333;
            background: red !important;
            border: none;
            border-radius: 8px;
            padding: 0;
            z-index: 99999 !important;
            box-shadow: none;
        }

        .attachment-preview img {
            width: 48px;
            height: 48px;
            object-fit: cover;
            border-radius: 6px;
            border: 1px solid #ddd;
            flex-shrink: 0;
        }

        .attachment-preview .file-name {
            flex: 1;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
            font-weight: 500;
            color: #333;
        }

        .attachment-preview .remove-attachment {
            background: none;
            border: none;
            color: #999;
            cursor: pointer;
            font-size: 18px;
            padding: 0;
            margin: 0;
            flex-shrink: 0;
        }

        .attachment-preview .remove-attachment:hover {
            color: #333;
        }

        .attachment-preview .file-icon {
            font-size: 24px;
        }

        .attach-button {
            position: absolute;
            right: 12px;
            top: 50%;
            transform: translateY(-50%);
            width: 24px;
            height: 24px;
            border: none;
            background: transparent;
            cursor: pointer;
            font-size: 18px;
            color: #007bff;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 0;
            z-index: 1002;
        }

        .attach-button::before {
            content: '📎';
        }

        .attach-options {
            position: absolute;
            right: 0;
            bottom: calc(100% + 8px);
            top: auto;
            transform: none;
            background: white;
            border: 1px solid #ddd;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.15);
            display: flex;
            flex-direction: column;
            gap: 3px;
            padding: 2px;
            z-index: 1003;
            min-width: 64px;
            max-width: 75px;
        }

        .attach-options button {
            border: none;
            background: transparent;
            cursor: pointer;
            padding: 5px 8px;
            border-radius: 6px;
            font-size: 13px;
        }

        .attach-options button:hover {
            background: #f0f0f0;
        }

        .camera-modal {
            position: fixed;
            inset: 0;
            background: rgba(0,0,0,0.6);
            display: flex;
            align-items: center;
            justify-content: center;
            z-index: 10000;
        }

        .camera-content {
            background: white;
            border-radius: 10px;
            padding: 16px;
            width: min(90vw, 420px);
            text-align: center;
        }

        .camera-content video {
            width: 100%;
            height: auto;
            border-radius: 8px;
            background: black;
        }

        .camera-buttons {
            display: flex;
            justify-content: space-between;
            margin-top: 10px;
            gap: 8px;
        }

        .camera-buttons button {
            flex: 1;
            padding: 8px 10px;
            border: none;
            border-radius: 6px;
            font-size: 14px;
            cursor: pointer;
        }

        .camera-buttons button:first-child {
            background: #007bff;
            color: #fff;
        }

        .camera-buttons button:last-child {
            background: #ccc;
            color: #000;
        }

        .send-button {
            position: absolute;
            right: 30px;
            top: 50%;
            transform: translateY(-50%);
            width: 36px;
            height: 36px;
            background: transparent;
            border: none;
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: center;
            transition: all 0.3s ease;
            flex-shrink: 0;
        }

        .send-button::before {
            content: '';
            position: absolute;
            width: 0;
            height: 0;
            border-left: 15px solid black;
            border-top: 12px solid transparent;
            border-bottom: 12px solid transparent;
            left: 14%;
            top: 50%;
            transform: translate(-50%, -50%);
        }

        .send-button.stop-typing::before {
            width: 18px;
            height: 18px;
            border: none;
            background: black;
        }

        .status {
            margin-top: 10px;
            font-size: 14px;
            color: #666;
        }
    </style>
</head>
<body>
    <div class="header">
    <img src="/logo?v=4" alt="OPUS" class="header-logo">
</div>
    <div class="chat-container" id="chat-container">
        <div class="message bot-message"><span class="message-prefix">ZIA:</span> <span>Olá! Sou ZIA, a inteligência artificial da Opus.</span></div>
    </div>

    <input type="file" id="hidden-file-input" accept=".pdf,.jpg,.jpeg,.xlsx,.xls" style="display:none;">

    <div id="camera-modal" class="camera-modal" style="display:none;">
        <div class="camera-content">
            <video id="camera-preview" autoplay playsinline></video>
            <div class="camera-buttons">
                <button id="capture-photo">Capturar</button>
                <button id="cancel-camera">Cancelar</button>
            </div>
            <canvas id="camera-canvas" style="display:none;"></canvas>
        </div>
    </div>

    <div class="input-container">
        <div class="input-wrapper">
            <div id="attachment-preview" class="attachment-preview"></div>
            <input type="text" id="user-input" placeholder="Digite sua pergunta..." onkeypress="handleKeyPress(event)">
            <div id="attach-options" class="attach-options" style="display:none;">
                <button type="button" id="camera-option">📷 Câmera</button>
                <button type="button" id="file-option">📁 Arquivo</button>
            </div>
            <button type="button" class="attach-button" id="attach-btn" title="Anexar arquivo"></button>
        </div>
        <button class="send-button"></button>
    </div>
    <div class="status" id="status"></div>

    <script>
        let currentContext = null;
let selectedPES = null;
let attachedFile = null;
let catalogoModulos = [];

        function setStatus(msg) {
            const el = document.getElementById('status');
            if (el) el.textContent = msg;
        }

        function updateAttachmentPreview() {
            const previewEl = document.getElementById('attachment-preview');
            const inputWrapper = document.querySelector('.input-wrapper');
            
            if (attachedFile) {
                console.log('attachedFile:', attachedFile);
                previewEl.style.opacity = '1';
                previewEl.style.pointerEvents = 'auto';
                inputWrapper.classList.add('has-attachment');
                previewEl.innerHTML = '';
                
                if (attachedFile.type === 'camera' || (attachedFile.data && attachedFile.data.startsWith('data:image/'))) {
                    // Para imagens, mostrar miniatura
                    const img = document.createElement('img');
                    img.src = attachedFile.data;
                    img.alt = attachedFile.name;
                    previewEl.appendChild(img);
                } else {
                    // Para arquivos, mostrar ícone específico
                    const icon = document.createElement('span');
                    icon.className = 'file-icon';
                    
                    if (attachedFile.name.endsWith('.pdf')) {
                        icon.textContent = '📄';
                    } else if (attachedFile.name.endsWith('.xlsx') || attachedFile.name.endsWith('.xls')) {
                        icon.textContent = '📊';
                    } else if (attachedFile.name.endsWith('.jpg') || attachedFile.name.endsWith('.jpeg')) {
                        icon.textContent = '🖼️';
                    } else {
                        icon.textContent = '📎';
                    }
                    previewEl.appendChild(icon);
                }
                
                // Adicionar nome do arquivo
                const nameEl = document.createElement('span');
                nameEl.className = 'file-name';
                nameEl.textContent = attachedFile.name;
                previewEl.appendChild(nameEl);
                
                // Adicionar botão de remover
                const removeBtn = document.createElement('button');
                removeBtn.className = 'remove-attachment';
                removeBtn.type = 'button';
                removeBtn.textContent = '✕';
                removeBtn.onclick = function() {
                    attachedFile = null;
                    updateAttachmentPreview();
                    setStatus('Anexo removido.');
                };
                previewEl.appendChild(removeBtn);
            } else {
                previewEl.style.opacity = '0';
                previewEl.style.pointerEvents = 'none';
                inputWrapper.classList.remove('has-attachment');
                previewEl.innerHTML = '';
            }
        }

        function renderOptions(options, onClick) {
            const chatContainer = document.getElementById('chat-container');
            
            // Remover container anterior se existir
            const oldContainer = document.getElementById('dynamic-action-buttons');
            if (oldContainer) oldContainer.remove();
            
            if (options.length === 0) return; // Se array vazio, não criar nada
            
            // Criar novo container para os botões
            const container = document.createElement('div');
            container.id = 'dynamic-action-buttons';
            container.className = 'context-selector';
            
            options.forEach(opt => {
                const btn = document.createElement('button');
                btn.className = 'context-btn';
                btn.textContent = opt.label;
                btn.onclick = () => { onClick(opt.value); };
                container.appendChild(btn);
            });
            
            // Adicionar ao chat
            chatContainer.appendChild(container);
        }

        async function carregarCatalogo() {
    try {
        const response = await fetch('/api/catalogo');
        const data = await response.json();

        if (data && data.sucesso && Array.isArray(data.modulos)) {
            catalogoModulos = data.modulos;
        }
    } catch (err) {
        console.error('Erro ao carregar catálogo:', err);
    }

    if (!catalogoModulos.length) {
        catalogoModulos = [
            {
                chave: 'orcamento',
                nome: 'Orçamento',
                submodulos: []
            },
            {
                chave: 'qualidade',
                nome: 'Qualidade',
                submodulos: []
            }
        ];
    }
}

function getModuloAtual() {
    return catalogoModulos.find(function(m) {
        return m.chave === currentContext;
    });
}

function moduloExigeSubmodulo() {
    const modulo = getModuloAtual();
    return modulo && modulo.submodulos && modulo.submodulos.length > 0;
}

function promptContext() {
    addMessageAnimated('Por favor, escolha um módulo:', false);

    renderOptions(catalogoModulos.map(function(modulo) {
        return {
            label: modulo.nome,
            value: modulo.chave
        };
    }), function(chaveModulo) {
        const modulo = catalogoModulos.find(function(m) {
            return m.chave === chaveModulo;
        });

        if (!modulo) {
            addMessageAnimated('Módulo não encontrado.', false);
            return;
        }

        currentContext = modulo.chave;
        selectedPES = null;

        addMessageAnimated('Modo selecionado: ' + modulo.nome, false);

        if (modulo.submodulos && modulo.submodulos.length > 0) {
            promptSubmodulos(modulo);
            setStatus(modulo.nome + ' selecionado: escolha uma opção.');
        } else {
            renderOptions([], function() {});
            setStatus(modulo.nome + ' ativo. Faça sua pergunta.');
        }
    });
}

function promptSubmodulos(modulo) {
    renderOptions(modulo.submodulos.map(function(sub) {
        return {
            label: sub.nome,
            value: sub.chave
        };
    }), function(chaveSubmodulo) {
        const sub = modulo.submodulos.find(function(s) {
            return s.chave === chaveSubmodulo;
        });

        const label = sub ? sub.nome : chaveSubmodulo;

        selectedPES = chaveSubmodulo;

        renderOptions([], function() {});
        addMessageAnimated('Tag selecionada: ' + label, false);
        setStatus(modulo.nome + ' - ' + label + '. Faça sua pergunta.');
    });
}

        // File handling (anexo via clip no prompt)
        const attachBtn = document.getElementById('attach-btn');
        const attachOptions = document.getElementById('attach-options');
        const fileInput = document.getElementById('hidden-file-input');
        const cameraModal = document.getElementById('camera-modal');
        const cameraPreview = document.getElementById('camera-preview');
        const cameraCanvas = document.getElementById('camera-canvas');
        const captureBtn = document.getElementById('capture-photo');
        const cancelCameraBtn = document.getElementById('cancel-camera');

        attachBtn.addEventListener('click', function() {
            attachOptions.style.display = attachOptions.style.display === 'flex' ? 'none' : 'flex';
        });

        document.getElementById('camera-option').addEventListener('click', async function() {
            console.log('camera-option clicado');
            attachOptions.style.display = 'none';

            if (!navigator.mediaDevices || !navigator.mediaDevices.getUserMedia) {
                setStatus('API de câmera não suportada neste navegador. Atualize o navegador.');
                return;
            }

            if (location.protocol !== 'https:' && location.hostname !== 'localhost' && location.hostname !== '127.0.0.1') {
                setStatus('Câmera exige HTTPS. Use https:// ou http://localhost.');
                return;
            }

            try {
                const constraints = {
                    video: {
                        facingMode: { ideal: 'environment' },
                        width: { ideal: 1280 },
                        height: { ideal: 720 }
                    },
                    audio: false
                };

                const stream = await navigator.mediaDevices.getUserMedia(constraints);
                console.log('Camera stream recebido', stream);
                cameraPreview.srcObject = stream;
                cameraPreview.play().catch(err => console.warn('cameraPreview play error', err));

                cameraModal.style.display = 'flex';
                cameraModal.setAttribute('data-streamed', 'true');
                setStatus('Câmera ativada. Clique em "Capturar" para tirar a foto.');
            } catch (err) {
                console.error('Camera error:', err);
                if (err.name === 'NotAllowedError' || err.name === 'PermissionDeniedError') {
                    setStatus('Permissão de câmera negada. Permita o acesso à câmera no navegador.');
                } else if (err.name === 'NotFoundError' || err.name === 'NotReadableError') {
                    setStatus('Câmera não encontrada ou ocupada por outro aplicativo.');
                } else if (err.name === 'NotSupportedError' || err.name === 'SecurityError') {
                    setStatus('Câmera exige HTTPS. Use um servidor HTTPS ou localhost.');
                } else {
                    setStatus('Erro ao acessar câmera: ' + (err.message || err));
                }
            }
        });

        document.getElementById('file-option').addEventListener('click', function() {
            attachOptions.style.display = 'none';
            fileInput.value = '';
            fileInput.click();
        });

        fileInput.addEventListener('click', () => {
            console.log('input clicado');
        });

        fileInput.addEventListener('change', function(e) {
            console.log('CHANGE DISPAROU');
            const file = e.target.files[0];
            if (file) {
                const allowedTypes = ['application/pdf', 'image/jpeg', 'image/jpg', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel'];
                if (!allowedTypes.includes(file.type)) {
                    setStatus('Formato não suportado. Use PDF, JPG, JPEG ou EXCEL.');
                    return;
                }

                const reader = new FileReader();
                reader.onload = function(ev) {
                    console.log('FILE LOADED');
                    attachedFile = {
                        type: 'file',
                        name: file.name,
                        data: ev.target.result
                    };
                    console.log(attachedFile);
                    setStatus(`Arquivo anexado: ${file.name}`);
                    updateAttachmentPreview();
                };
                reader.readAsDataURL(file);
            }
        });

        captureBtn.addEventListener('click', function() {
            const stream = cameraPreview.srcObject;
            if (!stream) {
                setStatus('Nenhum stream de câmera disponível. Tente abrir a câmera novamente.');
                return;
            }
            const width = cameraPreview.videoWidth;
            const height = cameraPreview.videoHeight;
            if (width === 0 || height === 0) {
                setStatus('Aguardando vídeo da câmera...');
                return;
            }
            cameraCanvas.width = width;
            cameraCanvas.height = height;
            const ctx = cameraCanvas.getContext('2d');
            ctx.drawImage(cameraPreview, 0, 0, width, height);
            const dataURL = cameraCanvas.toDataURL('image/jpeg', 0.92);
            attachedFile = { type: 'camera', name: 'camera_capture.jpg', data: dataURL };
            setStatus('Foto capturada e anexada para análise.');
            updateAttachmentPreview();
            stopCamera();
        });

        cancelCameraBtn.addEventListener('click', function() {
            stopCamera();
            setStatus('Captura de câmera cancelada.');
        });

        function stopCamera() {
            const stream = cameraPreview.srcObject;
            if (stream) {
                stream.getTracks().forEach(track => track.stop());
            }
            cameraPreview.srcObject = null;
            cameraModal.style.display = 'none';
            cameraModal.removeAttribute('data-streamed');
        }

        function addMessage(text, isUser = false) {
            const container = document.getElementById('chat-container');
            const messageDiv = document.createElement('div');
            messageDiv.className = `message ${isUser ? 'user-message' : 'bot-message'}`;
            const prefix = isUser ? 'Você: ' : 'ZIA: ';
            
            // Criar mensagem com prefixo colorido
            const prefixSpan = document.createElement('span');
            prefixSpan.className = 'message-prefix';
            prefixSpan.textContent = prefix;
            
            const textSpan = document.createElement('span');
            textSpan.textContent = text;
            
            messageDiv.appendChild(prefixSpan);
            messageDiv.appendChild(textSpan);
            
            container.appendChild(messageDiv);
            container.scrollTop = container.scrollHeight;
        }

        let isTyping = false;
        let typingTimeout = null;

        function addMessageAnimated(text, isUser = false, onComplete = null) {
            const container = document.getElementById('chat-container');
            const messageDiv = document.createElement('div');
            messageDiv.className = `message ${isUser ? 'user-message' : 'bot-message'}`;
            const prefix = isUser ? 'Você: ' : 'ZIA: ';
            
            // Criar mensagem com prefixo colorido
            const prefixSpan = document.createElement('span');
            prefixSpan.className = 'message-prefix';
            prefixSpan.textContent = prefix;
            
            const textSpan = document.createElement('span');
            textSpan.textContent = '';
            
            messageDiv.appendChild(prefixSpan);
            messageDiv.appendChild(textSpan);
            
            container.appendChild(messageDiv);
            
            // Animação letra por letra
            let index = 0;
            isTyping = true;
            updateSendButton();
            
            const typeWriter = () => {
                if (!isTyping) {
                    // Digitação interrompida
                    textSpan.textContent = text;
                    container.scrollTop = container.scrollHeight;
                    if (onComplete) onComplete();
                    return;
                }
                
                if (index < text.length) {
                    textSpan.textContent += text.charAt(index);
                    index++;
                    container.scrollTop = container.scrollHeight;
                    typingTimeout = setTimeout(typeWriter, 50); // 50ms por letra
                } else {
                    // Digitação completa
                    isTyping = false;
                    updateSendButton();
                    if (onComplete) onComplete();
                }
            };
            typeWriter();
        }

        function stopTyping() {
            if (isTyping) {
                isTyping = false;
                if (typingTimeout) {
                    clearTimeout(typingTimeout);
                    typingTimeout = null;
                }
                updateSendButton();
            }
        }

        function updateSendButton() {
            const button = document.querySelector('.send-button');
            if (!button) return;
            if (isTyping) {
                button.classList.add('stop-typing');
                button.onclick = stopTyping;
            } else {
                button.classList.remove('stop-typing');
                button.onclick = sendMessage;
            }
        }

        function addTypingIndicator() {
            const container = document.getElementById('chat-container');
            const typingDiv = document.createElement('div');
            typingDiv.className = 'message bot-message';
            typingDiv.id = 'typing-indicator';
            const prefixSpan = document.createElement('span');
            prefixSpan.className = 'message-prefix';
            prefixSpan.textContent = 'ZIA: ';
            const typingSpan = document.createElement('span');
            typingSpan.className = 'typing-indicator';
            typingDiv.appendChild(prefixSpan);
            typingDiv.appendChild(typingSpan);
            container.appendChild(typingDiv);
            container.scrollTop = container.scrollHeight;
        }

        function removeTypingIndicator() {
            const typingDiv = document.getElementById('typing-indicator');
            if (typingDiv) {
                typingDiv.remove();
            }
        }

        function sendMessage() {
            const input = document.getElementById('user-input');
            const message = input.value.trim();
            if (!message) return;

            // Construir mensagem do usuário incluindo anexo se houver
            let userMessage = message;
            if (attachedFile) {
                const attachmentText = `[Anexo: ${attachedFile.name || attachedFile.type}] `;
                userMessage = attachmentText + message;
            }

            addMessage(userMessage, true);
            input.value = '';

            if (!currentContext) {
                addMessageAnimated('Por favor, selecione primeiro Geral, Orçamento ou Qualidade.', false);
                return;
            }
            if (moduloExigeSubmodulo() && !selectedPES) {
    addMessageAnimated('Por favor, escolha uma opção antes de enviar a pergunta.', false);
    return;
}

            addTypingIndicator();

            // Compressar arquivo para mobile
            let fileToSend = attachedFile;
            if (attachedFile && attachedFile.type === 'camera') {
                // Reduzir qualidade de imagem da câmera em dispositivos móveis
                const canvas = document.createElement('canvas');
                const img = new Image();
                img.onload = function() {
                    canvas.width = img.width * 0.6;
                    canvas.height = img.height * 0.6;
                    const ctx = canvas.getContext('2d');
                    ctx.drawImage(img, 0, 0, canvas.width, canvas.height);
                    fileToSend.data = canvas.toDataURL('image/jpeg', 0.6);
                    sendAskRequest(userMessage, fileToSend);
                };
                img.src = attachedFile.data;
            } else {
                sendAskRequest(userMessage, fileToSend);
            }
        }

        function sendAskRequest(userMessage, fileToSend) {
            const controller = new AbortController();
            const timeout = setTimeout(() => controller.abort(), 30000); // 30 segundos timeout

            fetch('/ask', {
                method: 'POST',
                headers: {
                    'Content-Type': 'application/json',
                },
                body: JSON.stringify({ 
                    question: userMessage, 
                    context: currentContext,
                    pes: selectedPES,
                    file: fileToSend 
                }),
                signal: controller.signal
            })
            .then(response => {
                clearTimeout(timeout);
                if (!response.ok) {
                    throw new Error(`Erro HTTP: ${response.status}`);
                }
                return response.json();
            })
            .then(data => {
                removeTypingIndicator();
                if (data.response) {
                    addMessageAnimated(data.response);
                } else {
                    addMessageAnimated('Erro: resposta inválida do servidor.');
                }
                attachedFile = null;
                updateAttachmentPreview();
                setStatus('');
            })
            .catch(error => {
                clearTimeout(timeout);
                removeTypingIndicator();
                let mensagemErro = 'Erro ao processar a pergunta.';
                if (error.name === 'AbortError') {
                    mensagemErro = 'A requisição demorou muito. Verifique sua conexão de internet e tente novamente.';
                } else if (!navigator.onLine) {
                    mensagemErro = 'Sem conexão de internet. Verifique sua rede.';
                } else {
                    mensagemErro = `Erro: ${error.message || 'Falha ao conectar ao servidor'}`;
                    console.error('Erro detalhado:', error);
                }
                addMessageAnimated(mensagemErro);
                attachedFile = null;
                updateAttachmentPreview();
                setStatus('Tente novamente em alguns segundos.');
            });
        }

        function handleKeyPress(event) {
            if (event.key === 'Enter') {
                sendMessage();
            }
        }

        // Inicializar botão
        updateSendButton();

        // iniciar com prompt de modo
        console.log('Iniciando promptContext...');
        setTimeout(function() {
    carregarCatalogo().then(function() {
        promptContext();
    });
}, 500);
    </script>
</body>
</html>
"""

@app.route('/')
def home():
    return render_template_string(HTML_TEMPLATE)

# ===== ROTAS DE ADMINISTRAÇÃO =====

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            return redirect(url_for('admin_panel'))
        else:
            return render_template_string(ADMIN_LOGIN_TEMPLATE, erro='Usuário ou senha incorretos')
    return render_template_string(ADMIN_LOGIN_TEMPLATE)

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))

@app.route('/admin', methods=['GET'])
@login_required
def admin_panel():
    modulos = get_modulos()
    return render_template_string(ADMIN_PANEL_TEMPLATE, modulos=modulos)

@app.route('/admin/conteudos/<int:modulo_id>', methods=['GET'])
@login_required
def admin_conteudos(modulo_id):
    """Retorna lista de conteúdos de um módulo."""
    conteudos = get_conteudos_modulo(modulo_id)
    return jsonify({'conteudos': conteudos})

@app.route('/admin/dados/<int:conteudo_id>', methods=['GET'])
@login_required
def admin_dados_get(conteudo_id):
    """Retorna os dados de um conteúdo específico."""
    dados = get_dados_conteudo(conteudo_id)
    return jsonify({
        'texto': dados.get('texto', ''),
        'tag': dados.get('tag', ''),
        'arquivo': dados.get('arquivo', '')
    })

@app.route('/admin/dados/<int:conteudo_id>', methods=['POST'])
@login_required
def admin_dados_post(conteudo_id):
    """Atualiza os dados de um conteúdo."""
    data = request.get_json()
    texto = data.get('texto', '')
    tag = data.get('tag', '')
    
    sucesso_texto, mensagem_texto = atualizar_dados_conteudo(conteudo_id, texto)
    sucesso_tag, mensagem_tag = atualizar_tag_conteudo(conteudo_id, tag)
    sucesso = sucesso_texto and sucesso_tag
    mensagem = mensagem_texto if not sucesso_texto else mensagem_tag
    return jsonify({'sucesso': sucesso, 'mensagem': mensagem})

@app.route('/admin/arquivo/<int:conteudo_id>', methods=['POST'])
@login_required
def admin_arquivo_post(conteudo_id):
    remove = request.form.get('remove') == '1'
    if remove:
        dados = get_dados_conteudo(conteudo_id)
        if dados.get('arquivo'):
            caminho = os.path.join(UPLOAD_FOLDER, dados['arquivo'])
            if os.path.exists(caminho):
                os.remove(caminho)
        sucesso, mensagem = atualizar_arquivo_conteudo(conteudo_id, '')
        return jsonify({'sucesso': sucesso, 'mensagem': mensagem, 'arquivo': ''})

    arquivo = request.files.get('arquivo')
    if not arquivo or arquivo.filename == '':
        return jsonify({'sucesso': False, 'mensagem': 'Nenhum arquivo enviado.'})

    if not allowed_file(arquivo.filename):
        return jsonify({'sucesso': False, 'mensagem': 'Formato não permitido. Use XLSX, XLS ou CSV.'})

    filename = secure_filename(arquivo.filename)
    stored_name = f"conteudo_{conteudo_id}_{filename}"
    caminho = os.path.join(UPLOAD_FOLDER, stored_name)
    arquivo.save(caminho)

    sucesso, mensagem = atualizar_arquivo_conteudo(conteudo_id, stored_name)
    return jsonify({'sucesso': sucesso, 'mensagem': mensagem, 'arquivo': stored_name, 'nome': filename})

@app.route('/admin/criar-conteudo', methods=['POST'])
@login_required
def admin_criar_conteudo():
    """Cria um novo conteúdo em um módulo."""
    data = request.get_json()
    modulo_id = data.get('modulo_id')
    nome = data.get('nome', '').strip()
    descricao = data.get('descricao', '').strip()
    tag = data.get('tag', '').strip()
    
    if not modulo_id or not nome:
        return jsonify({'sucesso': False, 'mensagem': 'Módulo e nome são obrigatórios'})
    
    sucesso, mensagem = criar_conteudo(modulo_id, nome, descricao, tag)
    return jsonify({'sucesso': sucesso, 'mensagem': mensagem})

@app.route('/admin/deletar-conteudo/<int:conteudo_id>', methods=['POST'])
@login_required
def admin_deletar_conteudo_endpoint(conteudo_id):
    """Deleta um conteúdo."""
    sucesso, mensagem = deletar_conteudo(conteudo_id)
    return jsonify({'sucesso': sucesso, 'mensagem': mensagem})

@app.route('/status')
def status():
    return jsonify({'chatgpt_enabled': CHATGPT_HABILITADO})


def enviar_imagem_sem_cache(caminho, mimetype='image/png'):
    response = send_file(caminho, mimetype=mimetype)
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@app.route('/logo')
def logo():
    nomes_possiveis = [
        'OPUS.png',
        'OPUS.PNG',
        'Opus.png',
        'opus.png',
        'logo.png',
        'Logo.png'
    ]

    for nome in nomes_possiveis:
        logo_path = os.path.join(BASE_DIR, nome)
        if os.path.exists(logo_path):
            return enviar_imagem_sem_cache(logo_path, 'image/png')

    return 'Logo não encontrada. Envie o arquivo OPUS.png na mesma pasta do zia_web.py.', 404


@app.route('/favicon')
def favicon():
    nomes_possiveis = [
        'favicon.png',
        'FAVICON.png',
        'icone.png',
        'Icone.png'
    ]

    for nome in nomes_possiveis:
        favicon_path = os.path.join(BASE_DIR, nome)
        if os.path.exists(favicon_path):
            return enviar_imagem_sem_cache(favicon_path, 'image/png')

    return '', 404


def build_quality_combined_base():
    combined = []
    for key, base in knowledge_bases.items():
        if key.startswith('ai_base_qualidade_') and isinstance(base, list):
            combined.extend(base)
    return combined


@app.route('/ask', methods=['POST'])
def ask():
    try:
        # Adicionar timeout para requisições
        import signal
        def timeout_handler(signum, frame):
            raise TimeoutError("Requisição expirou")
        
        data = request.get_json(force=True, silent=True)
        if not data:
            return jsonify({'response': 'Erro: dados inválidos recebidos.'}), 400
            
        question = data.get('question', '').strip()
        context = data.get('context', 'geral').strip()
        pes = data.get('pes')
        file_info = data.get('file')
        
        # Validação básica
        if not question and not file_info:
            return jsonify({'response': 'Por favor, digite uma pergunta ou anexe um arquivo.'}), 400
        
        db_available = True  # Flag para saber se banco de dados está disponível
        
        # Se há arquivo ou imagem, incluir no texto para análise preliminar
        if file_info:
            file_name = file_info.get('name') if isinstance(file_info, dict) else str(file_info)
            source = file_info.get('type') if isinstance(file_info, dict) else 'arquivo'
            if question.strip() == '':
                return jsonify({'response': f"Recebido {file_name} ({source}). Por favor forneça a pergunta para análise de acordo com o PES."})
            question = f"[Anexo: {file_name} | Fonte: {source}] {question}"

        # **Tentar buscar primeiro no banco de dados com a nova estrutura**
        try:
            if context == 'qualidade':
                if pes:
                    # Buscar no banco de dados do PES específico
                    dados_conteudo = buscar_conteudo_modulo('qualidade', pes)
                    if dados_conteudo.get('sucesso') is not False and dados_conteudo.get('texto'):
                        response = gerar_resposta_com_fallback(question, base=dados_conteudo['texto'])
                        if response:
                            return jsonify({'response': clean_direct_response(response)})
                else:
                    matches = buscar_conteudos_modulo_por_query('qualidade', question)
                    if matches:
                        combined_text = '\n\n'.join([f"{nome}: {texto}" for nome, texto in matches])
                        response = gerar_resposta_com_fallback(question, base=combined_text)
                        if response:
                            return jsonify({'response': clean_direct_response(response)})
            elif context == 'orcamento':
                if pes and pes != 'orcamento':
                    dados_conteudo = buscar_conteudo_modulo('orcamento', pes)
                    if dados_conteudo.get('sucesso') is not False:
                        if dados_conteudo.get('arquivo'):
                            arquivo_path = os.path.join(UPLOAD_FOLDER, dados_conteudo['arquivo'])
                            resumo = load_spreadsheet_summary(arquivo_path)
                            if resumo:
                                response = gerar_resposta_com_fallback(question, base=resumo)
                                if response:
                                    return jsonify({'response': clean_direct_response(response)})
                        if dados_conteudo.get('texto'):
                            response = gerar_resposta_com_fallback(question, base=dados_conteudo['texto'])
                            if response:
                                return jsonify({'response': clean_direct_response(response)})
                else:
                    dados_conteudo = buscar_conteudo_modulo('orcamento', 'orcamento')
                    if dados_conteudo.get('sucesso') is not False and dados_conteudo.get('texto'):
                        response = gerar_resposta_com_fallback(question, base=dados_conteudo['texto'])
                        if response:
                            return jsonify({'response': clean_direct_response(response)})
            else:
                dados_conteudo = buscar_conteudo_modulo(context, context)
                if dados_conteudo.get('sucesso') is not False and dados_conteudo.get('texto'):
                    response = gerar_resposta_com_fallback(question, base=dados_conteudo['texto'])
                    if response:
                        return jsonify({'response': clean_direct_response(response)})
        except Exception as db_err:
            print(f"[AVISO] Erro ao acessar banco de dados: {db_err}")
            db_available = False
            # Continuar com fallback JSON

        # Se não encontrou no banco de dados, usar as bases JSON antigas como fallback
        # Select appropriate knowledge base
        pes_specific = False  # flag para saber se foi selecionado um PES específico
        if context == 'qualidade':
            if pes:
                pes_key = f'ai_base_qualidade_{pes}.json'
                base = knowledge_bases.get(pes_key)
                pes_specific = True
                if not base:
                    base = []
            else:
                base = build_quality_combined_base()
        elif context == 'orcamento':
            base_file = 'ai_base_orcamento.json'
            if pes and pes != 'orcamento':
                base_file = f'ai_base_orcamento_{pes}.json'
                pes_specific = True
            base = knowledge_bases.get(base_file, [])
        else:
            base_file = {
                'geral': 'ai_base_geral.json',
                'orcamento': 'ai_base_orcamento.json',
            }.get(context, 'ai_base_geral.json')
            base = knowledge_bases.get(base_file, {})

        # Tentar resposta direta da base de qualidade (PES) primeiro
        response = obter_resposta_base(question, base)
        if response:
            return jsonify({'response': clean_direct_response(response)})

        # Se um subcontexto específico foi selecionado e não encontrou resposta, não fazer fallback genérico
        if pes_specific and pes:
            if context == 'qualidade':
                question_for_fallback = f"[PES: {pes}] {question}"
            elif context == 'orcamento':
                question_for_fallback = f"[Orçamento: {pes}] {question}"
            else:
                question_for_fallback = question

            response = gerar_resposta_com_fallback(question_for_fallback, base=base)
            if response:
                return jsonify({'response': clean_direct_response(response)})
            else:
                contexto_texto = 'PES' if context == 'qualidade' else 'Orçamento'
                return jsonify({'response': f'Não encontrei informação sobre "{question}" no {contexto_texto} {pes}.'})
        
        # Forçar contexto na pergunta para fallback ChatGPT
        if context == 'qualidade' and pes:
            question_for_fallback = f"[PES: {pes}] {question}"
        elif context == 'orcamento' and pes and pes != 'orcamento':
            question_for_fallback = f"[Orçamento: {pes}] {question}"
        else:
            question_for_fallback = question

        # Use ChatGPT com fallback
        response = gerar_resposta_com_fallback(question_for_fallback, base=base)
        if response:
            return jsonify({'response': clean_direct_response(response)})
        else:
            return jsonify({'response': 'Não encontrei informação suficiente para responder. Tente reformular sua pergunta.'})
    
    except TimeoutError as te:
        error_msg = str(te)
        print(f"[TIMEOUT] Requisição expirou: {error_msg}")
        return jsonify({'response': 'A resposta demorou muito tempo. Verifique sua conexão de internet e tente novamente em alguns segundos.'}), 504
    
    except Exception as e:
        error_type = type(e).__name__
        error_msg = str(e)
        print(f"[ERRO] Exceção na rota /ask ({error_type}): {error_msg}")
        
        # Retornar mensagens de erro mais específicas
        if 'timeout' in error_msg.lower() or 'connection' in error_msg.lower():
            return jsonify({'response': 'Problema de conexão. Verifique sua internet e tente novamente.'}), 504
        else:
            return jsonify({'response': 'Desculpe, não consegui processar sua pergunta. Tente novamente em poucos instantes.'}), 500

@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    
    # Segurança e compatibilidade Cloudflare
    response.headers.add('Strict-Transport-Security', 'max-age=31536000; includeSubDomains; preload')
    response.headers.add('X-Forwarded-Proto', request.headers.get('X-Forwarded-Proto', 'http'))
    response.headers.add('X-Content-Type-Options', 'nosniff')
    response.headers.add('X-Frame-Options', 'DENY')
    response.headers.add('Referrer-Policy', 'strict-origin-when-cross-origin')
    response.headers.add('Cache-Control', 'public, max-age=3600')
    response.headers.add('Vary', 'Accept-Encoding')
    
    return response

if __name__ == '__main__':
    # Configuração para hospedagem livre (Render, Railway, etc)
    PORT = int(os.getenv('PORT', 5000))
    DEBUG = os.getenv('FLASK_DEBUG', 'False') == 'True'
    HOST = os.getenv('FLASK_HOST', '0.0.0.0')
    
    app.run(debug=DEBUG, host=HOST, port=PORT)
